"""
Azure Function: Update Avid RMS Master Tracker via Azure Table Storage.

Reads/writes go to Table Storage (instant point operations, no file I/O).
An /export-excel endpoint generates the Excel on demand from Table Storage.
"""

import azure.functions as func
import json
import logging
import os
import io
import traceback
from datetime import datetime, timezone
from azure.identity import DefaultAzureCredential
from azure.core.exceptions import ClientAuthenticationError, ResourceNotFoundError
from azure.data.tables import TableServiceClient, UpdateMode
from azure.storage.blob import BlobServiceClient, ContentSettings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = func.FunctionApp()

STORAGE_ACCOUNT = os.environ.get("STORAGE_ACCOUNT_NAME", "rgavidpoc950b")
TABLE_NAME = os.environ.get("TABLE_NAME", "rmstracker")
BLOB_CONTAINER = os.environ.get("BLOB_CONTAINER_NAME", "tracker")
BLOB_NAME = os.environ.get("BLOB_EXCEL_NAME", "Avid_RMS_Master_Tracker.xlsx")
MAX_FIELD_LENGTH = 5000

VALID_FIELDS = {
    "rms_number", "revision", "part_number", "pn_sub_code",
    "material_name", "material_class", "description", "intended_use",
    "approved_vendors", "catalog_numbers", "fill_volume", "container_closure",
    "storage_conditions", "expiration_retest", "sampling_instructions",
    "release_tests", "qualification_tests", "coa_requirements",
    "sop_references", "marking_packaging", "status", "notes",
}

# Full field order for detailed (agent-enriched) export
FIELD_ORDER = [
    "rms_number", "revision", "part_number", "pn_sub_code",
    "material_name", "material_class", "description", "intended_use",
    "approved_vendors", "catalog_numbers", "fill_volume", "container_closure",
    "storage_conditions", "expiration_retest", "sampling_instructions",
    "release_tests", "qualification_tests", "coa_requirements",
    "sop_references", "marking_packaging", "status", "last_updated", "notes",
]

EXCEL_HEADERS = [
    "RMS Number", "Revision", "Part Number (PN)", "PN-Sub Code",
    "Material Name", "Class", "Description", "Intended Use",
    "Approved Vendors", "Catalog Numbers", "Fill Volume", "Container Closure",
    "Storage Conditions", "Expiration / Retest", "Sampling Instructions",
    "Release Tests (Test | Method | Spec)", "Qualification Tests (Test | Method | Spec)",
    "CoA Requirements", "SOP / Handling References", "Marking & Packaging",
    "Status", "Last Updated", "Notes",
]



def sanitize_input(data: dict) -> dict:
    cleaned = {}
    for key, value in data.items():
        if key not in VALID_FIELDS:
            continue
        if not isinstance(value, str):
            value = str(value)
        cleaned[key] = value[:MAX_FIELD_LENGTH]
    return cleaned


def get_table_client():
    credential = DefaultAzureCredential()
    service = TableServiceClient(
        endpoint=f"https://{STORAGE_ACCOUNT}.table.core.windows.net",
        credential=credential,
    )
    return service.get_table_client(TABLE_NAME)


# --------------- UPDATE RMS ---------------

@app.function_name(name="UpdateRMSTracker")
@app.route(route="update-rms", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def update_rms(req: func.HttpRequest) -> func.HttpResponse:
    logging.info(f"update-rms called. Method={req.method}")

    try:
        body = req.get_json()
    except ValueError:
        return func.HttpResponse(
            json.dumps({"success": False, "error": "Invalid JSON body"}),
            status_code=200, mimetype="application/json",
        )

    body = sanitize_input(body)
    rms_number = body.get("rms_number")
    if not rms_number:
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"rms_number is required. Received fields: {list(body.keys())}"}),
            status_code=200, mimetype="application/json",
        )

    try:
        table_client = get_table_client()
        partition_key = "RMS"
        row_key = rms_number.strip().upper()

        # Check if entity already exists
        existing = None
        try:
            existing = table_client.get_entity(partition_key, row_key)
            action = "updated"
        except ResourceNotFoundError:
            action = "added"

        # Build entity
        entity = {
            "PartitionKey": partition_key,
            "RowKey": row_key,
        }
        for field in VALID_FIELDS:
            if field in body:
                entity[field] = body[field]

        if action == "updated":
            prev_notes = existing.get("notes", "") if existing else ""
            entity["notes"] = body.get("notes", prev_notes) + f" (Updated {datetime.now(timezone.utc).strftime('%Y-%m-%d')})"

        entity["status"] = body.get("status", "Populated")
        entity["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        if action == "added":
            table_client.create_entity(entity)
        else:
            table_client.update_entity(entity, mode=UpdateMode.MERGE)

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "action": action,
                "rms_number": rms_number,
                "message": f"Successfully {action} {rms_number}",
            }),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        tb = traceback.format_exc()
        logging.error(f"Auth failed for {rms_number}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Authentication failed (managed identity issue): {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        tb = traceback.format_exc()
        logging.error(f"Error updating RMS {rms_number}: {type(e).__name__}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- GET RMS ---------------

@app.function_name(name="GetRMSRow")
@app.route(route="get-rms", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def get_rms(req: func.HttpRequest) -> func.HttpResponse:
    logging.info(f"get-rms called. Params: {dict(req.params)}")

    rms_number = req.params.get("rms_number")
    if not rms_number:
        return func.HttpResponse(
            json.dumps({"found": False, "error": "rms_number query parameter is required"}),
            status_code=200, mimetype="application/json",
        )

    try:
        table_client = get_table_client()
        partition_key = "RMS"
        row_key = rms_number.strip().upper()

        try:
            entity = table_client.get_entity(partition_key, row_key)
        except ResourceNotFoundError:
            return func.HttpResponse(
                json.dumps({"found": False, "rms_number": rms_number}),
                status_code=200, mimetype="application/json",
            )

        row_data = {}
        for field in FIELD_ORDER:
            val = entity.get(field)
            if val is not None:
                row_data[field] = str(val)

        return func.HttpResponse(
            json.dumps({"found": True, "data": row_data}),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        tb = traceback.format_exc()
        logging.error(f"Auth failed reading {rms_number}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"found": False, "error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        tb = traceback.format_exc()
        logging.error(f"Error reading RMS {rms_number}: {type(e).__name__}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"found": False, "error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- EXPORT EXCEL ---------------

@app.function_name(name="ExportExcel")
@app.route(route="export-excel", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def export_excel(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("export-excel called")

    try:
        excel_bytes = build_excel_from_table()

        return func.HttpResponse(
            body=excel_bytes,
            status_code=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Avid_RMS_Master_Tracker.xlsx"},
        )

    except ClientAuthenticationError as e:
        tb = traceback.format_exc()
        logging.error(f"Auth failed for export: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        tb = traceback.format_exc()
        logging.error(f"Error exporting Excel: {type(e).__name__}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- EXPORT EXCEL LINK ---------------

@app.function_name(name="ExportExcelLink")
@app.route(route="export-excel-link", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def export_excel_link(req: func.HttpRequest) -> func.HttpResponse:
    """Return a JSON response with the download URL for the Excel tracker."""
    logging.info("export-excel-link called")
    host = req.headers.get("Host", "avid-agent-func-dyh4ahbwf5cvdha0.swedencentral-01.azurewebsites.net")
    scheme = "https"
    download_url = f"{scheme}://{host}/api/export-excel"
    return func.HttpResponse(
        json.dumps({
            "download_url": download_url,
            "filename": "Avid_RMS_Master_Tracker.xlsx",
            "message": "Click the download_url to download the Excel tracker.",
        }),
        status_code=200, mimetype="application/json",
    )


# --------------- TIMER: SYNC TABLE → BLOB EXCEL ---------------

def build_excel_from_table() -> bytes:
    """Query all entities from Table Storage and return an Excel workbook
    with all 23 columns matching the full RMS Master Tracker format."""
    table_client = get_table_client()
    entities = sorted(
        table_client.list_entities(),
        key=lambda e: e.get("RowKey", ""),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, entity in enumerate(entities, 2):
        for col_idx, field in enumerate(FIELD_ORDER, 1):
            val = entity.get(field)
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=str(val))

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


@app.function_name(name="SyncExcelToBlob")
@app.timer_trigger(schedule="0 */5 * * * *", arg_name="timer", run_on_startup=False)
def sync_excel_to_blob(timer: func.TimerRequest) -> None:
    """Every 5 minutes, export Table Storage to an Excel blob."""
    logging.info("SyncExcelToBlob triggered")
    try:
        excel_bytes = build_excel_from_table()

        credential = DefaultAzureCredential()
        blob_service = BlobServiceClient(
            account_url=f"https://{STORAGE_ACCOUNT}.blob.core.windows.net",
            credential=credential,
        )
        blob_client = blob_service.get_blob_client(BLOB_CONTAINER, BLOB_NAME)
        blob_client.upload_blob(
            excel_bytes,
            overwrite=True,
            content_settings=ContentSettings(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        )
        logging.info(f"Synced {BLOB_NAME} to container {BLOB_CONTAINER}")
    except Exception as e:
        logging.error(f"SyncExcelToBlob failed: {type(e).__name__}: {e}\n{traceback.format_exc()}")


# --------------- UPLOAD TRACKER (raw file → blob) ---------------

@app.function_name(name="UploadTracker")
@app.route(route="upload-tracker", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def upload_tracker(req: func.HttpRequest) -> func.HttpResponse:
    """Upload the customer's Excel tracker file directly to blob storage."""
    logging.info("upload-tracker called")

    try:
        file_bytes = req.get_body()
        if not file_bytes:
            return func.HttpResponse(
                json.dumps({"success": False, "error": "Empty request body — send the Excel file as binary body"}),
                status_code=200, mimetype="application/json",
            )

        credential = DefaultAzureCredential()
        blob_service = BlobServiceClient(
            account_url=f"https://{STORAGE_ACCOUNT}.blob.core.windows.net",
            credential=credential,
        )
        blob_client = blob_service.get_blob_client(BLOB_CONTAINER, BLOB_NAME)
        blob_client.upload_blob(
            file_bytes,
            overwrite=True,
            content_settings=ContentSettings(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": f"Uploaded tracker to {BLOB_CONTAINER}/{BLOB_NAME}",
                "size_bytes": len(file_bytes),
            }),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        tb = traceback.format_exc()
        logging.error(f"Auth failed for upload: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        tb = traceback.format_exc()
        logging.error(f"Error uploading tracker: {type(e).__name__}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- IMPORT EXCEL → Table Storage ---------------

@app.function_name(name="ImportExcel")
@app.route(route="import-excel", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def import_excel(req: func.HttpRequest) -> func.HttpResponse:
    """Parse the full 23-column Excel tracker and bulk-load rows into Table Storage.

    Accepts the Excel file as the raw binary request body, OR reads from blob
    storage when the body is empty or contains JSON {"source": "blob"}.
    """
    logging.info("import-excel called")

    try:
        body = req.get_body()
        from_blob = False

        # Determine source: uploaded bytes or read from blob
        if not body or len(body) < 100:
            # Try to interpret as JSON instruction
            try:
                payload = json.loads(body) if body else {}
            except (ValueError, json.JSONDecodeError):
                payload = {}
            if payload.get("source") == "blob" or not body:
                from_blob = True

        if from_blob:
            credential = DefaultAzureCredential()
            blob_service = BlobServiceClient(
                account_url=f"https://{STORAGE_ACCOUNT}.blob.core.windows.net",
                credential=credential,
            )
            blob_client = blob_service.get_blob_client(BLOB_CONTAINER, BLOB_NAME)
            body = blob_client.download_blob().readall()
            logging.info("import-excel: reading from blob storage")

        wb = load_workbook(filename=io.BytesIO(body), read_only=True, data_only=True)
        ws = wb.active

        table_client = get_table_client()
        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                skipped += 1
                continue

            rms_raw = str(row[0]).strip()
            if not rms_raw.upper().startswith("RMS"):
                skipped += 1
                continue

            row_key = rms_raw.upper().replace(" ", "")
            entity = {
                "PartitionKey": "RMS",
                "RowKey": row_key,
                "rms_number": rms_raw,
                "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "status": "Imported",
            }

            # Map all columns to table fields
            for col_idx, field in enumerate(FIELD_ORDER):
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val:
                        entity[field] = val[:MAX_FIELD_LENGTH]

            try:
                table_client.upsert_entity(entity, mode=UpdateMode.MERGE)
                imported += 1
            except Exception as row_err:
                errors.append({"row": row_idx, "rms": rms_raw, "error": str(row_err)[:200]})

        wb.close()

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "imported": imported,
                "skipped": skipped,
                "errors": errors,
                "source": "blob" if from_blob else "upload",
                "message": f"Imported {imported} rows into Table Storage",
            }),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        tb = traceback.format_exc()
        logging.error(f"Auth failed for import: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        tb = traceback.format_exc()
        logging.error(f"Error importing Excel: {type(e).__name__}: {e}\n{tb}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- HEALTH CHECK ---------------

@app.function_name(name="HealthCheck")
@app.route(route="health", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def health_check(req: func.HttpRequest) -> func.HttpResponse:
    config_status = {
        "STORAGE_ACCOUNT_NAME": bool(STORAGE_ACCOUNT),
        "TABLE_NAME": bool(TABLE_NAME),
    }
    auth_ok = False
    auth_error = None
    try:
        table_client = get_table_client()
        list(table_client.query_entities("PartitionKey eq 'RMS'", results_per_page=1))
        auth_ok = True
    except Exception as e:
        auth_error = f"{type(e).__name__}: {str(e)[:300]}"

    return func.HttpResponse(
        json.dumps({
            "status": "ok" if all(config_status.values()) and auth_ok else "error",
            "config": config_status,
            "auth": {"ok": auth_ok, "error": auth_error},
        }),
        status_code=200, mimetype="application/json",
    )


# --------------- helpers for blob listing ---------------

RMS_SOURCE_CONTAINER = os.environ.get("RMS_SOURCE_CONTAINER", "avid-rms")


def get_blob_service():
    credential = DefaultAzureCredential()
    return BlobServiceClient(
        account_url=f"https://{STORAGE_ACCOUNT}.blob.core.windows.net",
        credential=credential,
    )


def list_rms_blobs():
    """Return a sorted list of .docx blob names from the RMS source container."""
    blob_service = get_blob_service()
    container_client = blob_service.get_container_client(RMS_SOURCE_CONTAINER)
    blobs = []
    for blob in container_client.list_blobs():
        if blob.name.lower().endswith(".docx"):
            blobs.append(blob.name)
    return sorted(blobs)


def list_tracker_entries():
    """Return all RMS entities from Table Storage."""
    table_client = get_table_client()
    return list(table_client.list_entities())


# --------------- LIST RMS SOURCES ---------------

@app.function_name(name="ListRmsSources")
@app.route(route="list-rms-sources", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def list_rms_sources(req: func.HttpRequest) -> func.HttpResponse:
    """List all RMS .docx source files in blob storage."""
    logging.info("list-rms-sources called")
    try:
        blobs = list_rms_blobs()
        return func.HttpResponse(
            json.dumps({"total": len(blobs), "files": blobs}),
            status_code=200, mimetype="application/json",
        )
    except ClientAuthenticationError as e:
        logging.error(f"Auth failed for list-rms-sources: {e}")
        return func.HttpResponse(
            json.dumps({"error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        logging.error(f"Error listing RMS sources: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return func.HttpResponse(
            json.dumps({"error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- LIST RMS (all tracker entries) ---------------

@app.function_name(name="ListRms")
@app.route(route="list-rms", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def list_rms(req: func.HttpRequest) -> func.HttpResponse:
    """Return all RMS entries currently in Table Storage with optional status filter."""
    logging.info("list-rms called")
    status_filter = req.params.get("status")  # optional filter

    try:
        entities = list_tracker_entries()
        entries = []
        for entity in sorted(entities, key=lambda e: e.get("RowKey", "")):
            row = {}
            for field in FIELD_ORDER:
                val = entity.get(field)
                if val is not None:
                    row[field] = str(val)
            if status_filter and row.get("status", "").lower() != status_filter.lower():
                continue
            entries.append(row)

        return func.HttpResponse(
            json.dumps({"total": len(entries), "entries": entries}),
            status_code=200, mimetype="application/json",
        )
    except ClientAuthenticationError as e:
        logging.error(f"Auth failed for list-rms: {e}")
        return func.HttpResponse(
            json.dumps({"error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        logging.error(f"Error listing RMS: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return func.HttpResponse(
            json.dumps({"error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- BULK UPDATE RMS ---------------

@app.function_name(name="BulkUpdateRms")
@app.route(route="bulk-update-rms", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def bulk_update_rms(req: func.HttpRequest) -> func.HttpResponse:
    """Batch update/add multiple RMS rows. Expects JSON: {"entries": [...]}.
    Each entry follows the same schema as /update-rms."""
    logging.info("bulk-update-rms called")

    try:
        body = req.get_json()
    except ValueError:
        return func.HttpResponse(
            json.dumps({"success": False, "error": "Invalid JSON body"}),
            status_code=200, mimetype="application/json",
        )

    entries = body.get("entries")
    if not entries or not isinstance(entries, list):
        return func.HttpResponse(
            json.dumps({"success": False, "error": "'entries' array is required"}),
            status_code=200, mimetype="application/json",
        )

    if len(entries) > 200:
        return func.HttpResponse(
            json.dumps({"success": False, "error": "Maximum 200 entries per batch"}),
            status_code=200, mimetype="application/json",
        )

    try:
        table_client = get_table_client()
        results = {"added": 0, "updated": 0, "errors": []}

        for idx, raw_entry in enumerate(entries):
            entry = sanitize_input(raw_entry)
            rms_number = entry.get("rms_number")
            if not rms_number:
                results["errors"].append({"index": idx, "error": "missing rms_number"})
                continue

            partition_key = "RMS"
            row_key = rms_number.strip().upper()

            try:
                table_client.get_entity(partition_key, row_key)
                action = "updated"
            except ResourceNotFoundError:
                action = "added"

            entity = {"PartitionKey": partition_key, "RowKey": row_key}
            for field in VALID_FIELDS:
                if field in entry:
                    entity[field] = entry[field]
            entity["status"] = entry.get("status", "Populated")
            entity["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")

            try:
                table_client.upsert_entity(entity, mode=UpdateMode.MERGE)
                results[action] += 1
            except Exception as row_err:
                results["errors"].append({"index": idx, "rms_number": rms_number, "error": str(row_err)[:200]})

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "added": results["added"],
                "updated": results["updated"],
                "errors": results["errors"],
                "message": f"Processed {results['added'] + results['updated']} entries",
            }),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        logging.error(f"Auth failed for bulk-update-rms: {e}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        logging.error(f"Error in bulk-update-rms: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return func.HttpResponse(
            json.dumps({"success": False, "error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- TRACKER STATUS ---------------

@app.function_name(name="TrackerStatus")
@app.route(route="tracker-status", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def tracker_status(req: func.HttpRequest) -> func.HttpResponse:
    """Dashboard summary: source doc count, populated count, pending, last update."""
    logging.info("tracker-status called")

    try:
        # Count source blobs
        try:
            source_files = list_rms_blobs()
            source_count = len(source_files)
        except Exception as blob_err:
            logging.warning(f"Could not list source blobs: {blob_err}")
            source_files = []
            source_count = None

        # Count tracker entries
        entities = list_tracker_entries()
        tracker_entries = []
        status_counts = {}
        last_updated = None
        for entity in entities:
            rms = entity.get("rms_number") or entity.get("RowKey", "")
            tracker_entries.append(rms.upper().replace(" ", ""))
            status = entity.get("status", "Unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
            lu = entity.get("last_updated")
            if lu and (last_updated is None or str(lu) > str(last_updated)):
                last_updated = str(lu)

        populated = len(tracker_entries)

        # Derive pending count from source blobs if available
        if source_count is not None:
            source_rms_keys = set()
            for f in source_files:
                name = f.rsplit(".", 1)[0].upper().replace(" ", "").replace("_", "-")
                source_rms_keys.add(name)
            pending = len(source_rms_keys - set(tracker_entries))
        else:
            pending = None

        return func.HttpResponse(
            json.dumps({
                "source_docs": source_count,
                "populated": populated,
                "pending": pending,
                "status_breakdown": status_counts,
                "last_updated": last_updated,
            }),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        logging.error(f"Auth failed for tracker-status: {e}")
        return func.HttpResponse(
            json.dumps({"error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        logging.error(f"Error in tracker-status: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return func.HttpResponse(
            json.dumps({"error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )


# --------------- DIFF RMS ---------------

@app.function_name(name="DiffRms")
@app.route(route="diff-rms", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def diff_rms(req: func.HttpRequest) -> func.HttpResponse:
    """Compare source blobs vs tracker entries. Returns which RMS numbers are
    missing from the tracker, which are in the tracker but have no source file,
    and which are fully populated."""
    logging.info("diff-rms called")

    try:
        # Source blob names → RMS keys
        source_files = list_rms_blobs()
        source_map = {}
        for f in source_files:
            key = f.rsplit(".", 1)[0].upper().replace(" ", "").replace("_", "-")
            source_map[key] = f

        # Tracker keys
        entities = list_tracker_entries()
        tracker_keys = set()
        for entity in entities:
            rms = entity.get("rms_number") or entity.get("RowKey", "")
            tracker_keys.add(rms.upper().replace(" ", ""))

        source_keys = set(source_map.keys())
        missing_from_tracker = sorted(source_keys - tracker_keys)
        in_tracker_no_source = sorted(tracker_keys - source_keys)
        fully_populated = sorted(source_keys & tracker_keys)

        return func.HttpResponse(
            json.dumps({
                "missing_from_tracker": missing_from_tracker,
                "in_tracker_no_source": in_tracker_no_source,
                "fully_populated": fully_populated,
                "summary": {
                    "total_source": len(source_keys),
                    "total_in_tracker": len(tracker_keys),
                    "missing": len(missing_from_tracker),
                    "no_source": len(in_tracker_no_source),
                    "matched": len(fully_populated),
                },
            }),
            status_code=200, mimetype="application/json",
        )

    except ClientAuthenticationError as e:
        logging.error(f"Auth failed for diff-rms: {e}")
        return func.HttpResponse(
            json.dumps({"error": f"Authentication failed: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )
    except Exception as e:
        logging.error(f"Error in diff-rms: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return func.HttpResponse(
            json.dumps({"error": f"Internal error: {type(e).__name__}: {str(e)[:500]}"}),
            status_code=200, mimetype="application/json",
        )